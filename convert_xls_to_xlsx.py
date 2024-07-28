import os
import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
from itertools import cycle


def convert_xls_to_xlsx(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.xls'):
            file_path = os.path.join(directory, filename)
            try:
                print(f"Converting {filename}...")
                convert_file(file_path)
                print(f"Successfully converted {filename}.")
            except Exception as e:
                print(f"Failed to convert {filename}: {e}")


def convert_file(file_path):
    xls_wb = xlrd.open_workbook(file_path)
    new_wb = openpyxl.Workbook()
    sheets = xls_wb.sheets()

    for index, sheet in enumerate(sheets):
        if index == 0:
            new_ws = new_wb.active
            new_ws.title = sheet.name
        else:
            new_ws = new_wb.create_sheet(sheet.name)

        for rowIndex in range(sheet.nrows):
            row_values = sheet.row_values(rowIndex)
            row_values = convert_date_cells(
                row_values, sheet, rowIndex, xls_wb.datemode)
            for colIndex, value in enumerate(row_values):
                new_ws.cell(rowIndex + 1, colIndex + 1).value = value

    save_converted_file(new_wb, file_path)
    style_workbook(file_path[:-4] + '.xlsx')


def convert_date_cells(row_values, sheet, rowIndex, datemode):
    for colIndex, _ in enumerate(row_values):
        if sheet.cell_type(rowIndex, colIndex) == xlrd.XL_CELL_DATE:
            date_value = sheet.cell_value(rowIndex, colIndex)
            try:
                date = xlrd.xldate.xldate_as_datetime(date_value, datemode)
                formatted_date = date.strftime("%d/%m/%Y")
                row_values[colIndex] = formatted_date
            except xlrd.xldate.XLDateError as e:
                print(
                    f"Error converting date at row {rowIndex+1}, column {colIndex+1}: {e}")
    return row_values


def save_converted_file(new_wb, original_file_path):
    new_file_path = original_file_path[:-4] + '.xlsx'
    new_wb.save(new_file_path)
    print(f"Saved converted file to {new_file_path}")


def style_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Set header style
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid")
        header_border = Border(left=Side(style='thin', color='FFFFFF'),
                               right=Side(style='thin', color='FFFFFF'),
                               top=Side(style='thin', color='FFFFFF'),
                               bottom=Side(style='thin', color='FFFFFF'))
        alignment = Alignment(horizontal="center",
                              vertical="center", wrap_text=True)

        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = alignment

        ws.row_dimensions[1].height = 40  # Height for header row

        # Auto-fit columns and rows (approximation)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 10)
            ws.column_dimensions[column].width = adjusted_width

        # Apply striped pattern to content rows
        odd_fill = PatternFill(start_color="B5E6A2",
                               end_color="B5E6A2", fill_type="solid")
        even_fill = PatternFill(start_color="DAF2D0",
                                end_color="DAF2D0", fill_type="solid")

        for row in range(2, max_row + 1):
            fill = even_fill if row % 2 == 0 else odd_fill
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = header_border

        # Freeze the top row
        ws.freeze_panes = ws['A2']

        # Add table-like appearance
        ws.auto_filter.ref = ws.dimensions

        # Special styling for specific workbook
        # if os.path.basename(file_path) == "Report_Bibendum_MatthewClark.xlsx":
            # style_groupage_no_column(ws)

    wb.save(file_path)
    print(f"Styled and saved workbook {file_path}")


def style_groupage_no_column(ws):
    # Locate the "Groupage No." column
    groupage_col_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Groupage No.":
            groupage_col_index = col
            break

    if groupage_col_index is None:
        print(f"'Groupage No.' column not found in worksheet {ws.title}")
        return

    # Identify unique values and assign colors
    value_frequency_mapping = defaultdict(int)
    value_color_mapping = defaultdict(lambda: None)
    color_palette = cycle([
        "974706FFFFFF", "215967FFFFFF", "403151FFFFFF", "4f6228FFFFFF", "632523FFFFFF", "244062FFFFFF", "0f243eFFFFFF",
        "1d1b10FFFFFF", "808080000000", "e26b0a000000", "31869bFFFFFF", "60497aFFFFFF", "76933c000000", "963634FFFFFF",
        '366092FFFFFF',
        '16365cFFFFFF',
        '494529FFFFFF',
        '262626FFFFFF',
        'a6a6a6000000',
        'fabf8f000000',
        '92cddc000000',
        'b1a0c7000000',
        'da9694000000',
        '95b3d7000000',
        '538dd5000000',
        '948a54000000',
        '404040FFFFFF',
        'fcd5b4000000',
        'b7dee8000000',
        'ccc0da000000',
        'e6b8b7000000',
        'b8cce4000000',
        '8db4e2000000',
        'c4bd97000000',
        '002060FFFFFF',
        '0070c0FFFFFF',
        '00b0f0000000',
        '00b050000000',
        'ffff00000000',
        'ffc000000000',
        'ff0000FFFFFF',
        'c00000FFFFFF',
    ])

    # Collecting all values in the column
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row, groupage_col_index).value
        if cell_value:
            value_frequency_mapping[cell_value] += 1
    # Assigning colors to each unique value
    for value in value_frequency_mapping:
        if (value_frequency_mapping[value]):
            value_color_mapping[value] = next(color_palette)

    # Apply colors to cells
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row, groupage_col_index).value
        if cell_value and cell_value in value_frequency_mapping:
            if value_frequency_mapping[cell_value] == 1:
                fill_color = '7030a0'
                ws.cell(row, groupage_col_index).font = Font(
                    color="FFFFFF", bold=True)
                ws.cell(row, groupage_col_index).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
            else:
                fill_color = value_color_mapping[cell_value][0:6]
                font_color = value_color_mapping[cell_value][6:12]
                # print(type(font_color))
                # print(font_color)
                ws.cell(row, groupage_col_index).font = Font(
                    color=font_color, bold=True)
                ws.cell(row, groupage_col_index).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")


# Specify the directory containing your xls files
directory = 'C:/Wineflow'
convert_xls_to_xlsx(directory)
